# This home-made modules contains functions to import data from various sources.
# In order to get Python to see the module, added following to .bash_profile or .bashrc:
#
# PYTHONPATH=$PYTHONPATH:/path/to/folder; export PYTHONPATH
#
# To use the functions in this module, import phjReadData and refer to functions using:
# df = phjReadData.function_name(arg)
#
# e.g.
# df = phjReadData.phjReadDataFromExcelNamedCellRange('/Users/phil/Dropbox/Clarkson/phjDataweatheranalysis.xlsx', '2011 data', 'tides_data')


#######################################################
# Read data from a named cell range in Excel workbook #
#######################################################
def phjReadDataFromExcelNamedCellRange( phjExcelFileName,
										phjExcelRangeName,
										phjDatetimeFormat = "%Y-%m-%d %H:%M:%S",
										phjMissingValue = "missing",
										phjHeaderRow = False,
										phjPrintResults = False):
										
	# This function reads data from a named cell range in an Excel spreadsheet (.xlsx).
	# Input phjExcelFileName (including path to file), phjExcelSheetName and phjExcelRangeName.
	# Also input phjDatetimeFormat and phjPrintResults (default = False).
	# If the function can find the correct cell range, it reads the data and returns
	# it in the form of a pandas DataFrame.
	
	# IMPORT REQUIRED PACKAGES
	# ========================
	# The following method to check that a package is available is not considered to be
	# good practice:
	#
	#	try:
	# 		import package_name
	# 		HAVE_PACKAGE = True
	# 	except ImportError:
	# 		HAVE_PACKAGE = False
	#
	# Instead use the following method (as described at: http://developer.plone.org/reference_manuals/external/plone.api/contribute/conventions.html#about-imports)...
	
	import pkg_resources
	
	try:
		pkg_resources.get_distribution('openpyxl')
	except pkg_resources.DistributionNotFound:
		print("Error: openpyxl package not available.")
		return False
	else:
		import openpyxl
	
	
	try:
		pkg_resources.get_distribution('numpy')
	except pkg_resources.DistributionNotFound:
		print("Error: numpy package not available.")
		return False
	else:
		import numpy as np


	try:
		pkg_resources.get_distribution('pandas')
	except pkg_resources.DistributionNotFound:
		print("Error: pandas package not available.")
		return False
	else:
		import pandas as pd
	
	
	import re
	
	# LOAD DATA FROM EXCEL WORKBOOK
	# =============================
	try:
		# Load phjTempWorkbook. If unsuccessful, throws an InvalidFileException (imported in header)
		phjTempWorkbook = openpyxl.load_workbook(filename = phjExcelFileName, read_only = True, data_only = True)
	except FileNotFoundError:
		print("File named '" + phjExcelFileName + "' does not exist.")
		return None
	
	
	# Get named range of cells. Function returns None if named_range not found...
	phjTempCellRange = phjTempWorkbook.get_named_range(phjExcelRangeName)
		
	if phjTempCellRange == None:
		print("Cell range named '" + phjExcelRangeName + "' not found in phjTempWorkbook.")
		return None
	
	
	# Get name of worksheet from cell range instance
	# The cellrange.destinations is given as a list of tuples of the form:
	#
	#     [(<ReadOnlyWorksheet "Sheet1">, '$A$1:$D$100')]
	#
	# The first item in the first tuple is found using cellrange.destinations[0][0].
	# This needs to be converted to a string and a regular expression used to find the
	# text between the double quotation marks. The result is returned as a list (although,
	# in this case, there is only a single item). The first item is returned and
	# converted to a string. This can then be used to get a Worksheet instance.
	phjTempWorksheetName = str(re.findall(r'\"(.+?)\"',str(phjTempCellRange.destinations[0][0]))[0])
	phjTempWorksheet = phjTempWorkbook[phjTempWorksheetName]
	
	
	if phjPrintResults == True:
		print("\nList of iterable properties:")
		print(dir(phjTempCellRange))						# Get list of all iterable properties of object
		
		print("\nList of tuples of named ranges and cell ranges")
		print(phjTempCellRange.destinations)				# This gives a list containing tuples of worksheet names and cell ranges 
		
		print("\nFirst tuple of named ranges and cell range")
		print(phjTempCellRange.destinations[0])				# Give first tuple
		
		print("\nCell range")
		print(phjTempCellRange.destinations[0][1])			# Gives element [1] of tuple [0] i.e. the cell range

		print("\nLocal Sheet ID")
		print(phjTempCellRange.destinations[0][0])			# Gives element [1] of tuple [0] i.e. the cell range

	
	# Define temporary list to store data...
	phjTempImportedData = []
	
	# Step through each row in cell range.
	# (N.B. Cells returned by iter_rows() are not regular openpyxl.cell.cell.Cell but openpyxl.cell.read_only.ReadOnlyCell.)
	for phjTempRow in phjTempWorksheet.iter_rows(phjTempCellRange.destinations[0][1]):
	
		# Define temporary list to store values from phjTempCells in phjTempRows
		phjTempData=[]
		
		# Step through each phjTempCell in phjTempRow...
		for phjTempCell in phjTempRow:
			if phjTempCell.value == None:
				phjTempData.append(phjMissingValue)
			
			else:
				if phjTempCell.is_date:
					# If the phjTempCell contains a date, the format of phjTempCell.value is,
					# for example, datetime.datetime(2011, 1, 1, 0, 0). Therefore, reformat
					# using required format.
					phjTempData.append(phjTempCell.value.strftime(phjDatetimeFormat))
			
				elif phjTempCell.data_type == 's':				# TYPE_STRING = 's' AND TYPE_STRING_NULL = 's'
					phjTempData.append(phjTempCell.value)
			
				elif phjTempCell.data_type == 'f':				# TYPE_FORMULA = 'f'
					# Including 'data_only=True' in openpyxl.load_phjTempWorkbook() means that formulae aren't recognised as formulae, only the resulting value.
					phjTempData.append(phjTempCell.value)
			
				elif phjTempCell.data_type == 'n':				# TYPE_NUMERIC = 'n'
					# phjTempData.append(Decimal(phjTempCell.internal_value).quantize(Decimal('1.00')))
					phjTempData.append(phjTempCell.value)
					
				elif phjTempCell.data_type == 'b':				# TYPE_BOOL = 'b'
					phjTempData.append(phjTempCell.value)
			
				elif phjTempCell.data_type == 'inlineStr':		# TYPE_INLINE = 'inlineStr'
					phjTempData.append(phjTempCell.value)
				
				elif phjTempCell.data_type == 'e':				# TYPE_ERROR = 'e'
					phjTempData.append(phjTempCell.value)
			
				elif phjTempCell.data_type == 'str':			# TYPE_FORMULA_CACHE_STRING = 'str'
					phjTempData.append(phjTempCell.value)
			
				else:
					phjTempData.append(phjTempCell.value)
				
		phjTempImportedData.append(tuple(phjTempData))
		
	phjTempVariableNames = phjDealWithHeaderRow(phjTempImportedData,
												phjHeaderRow = phjHeaderRow,
												phjPrintResults = phjPrintResults)
	
	# Convert dataset to pandas DataFrame.
	# Each column now headed with original column headers as seen in Excel file
	# if header row present or with generic labels of 'var1', 'var2', etc. if no
	# header row present.
	
	phjTempDF = pd.DataFrame(phjTempImportedData, columns=phjTempVariableNames)
	
	if phjPrintResults == True:
		print("\nImported data")
		print("-------------")
		print(phjTempDF)

	return phjTempDF




def phjDealWithHeaderRow(phjData,
						 phjHeaderRow = False,
						 phjPrintResults = False):
	# This function gets the variable names from the first row of data and
	# removes the header row from the data list. The data is passed by
	# references and, therefore, it can be mutated without having to make
	# a copy of the whole dataset.
	
	# Deal with headers in first phjTempRow...
	if phjHeaderRow:
		# Identify variable names from first phjTempRow of data...
		phjTempVariableNames = phjData[0]
		
		# Remove header names from data...
		del phjData[0]
		
		if phjPrintResults == True:
			print("\nFirst row (containing variable names) has been removed from the data.")
		
	else:
		# If first row doesn't contain variable names then create list of names var1, var2, etc.
		phjTempVariableNames = []
		for i in range (len(phjData[0])):
			phjTempVariableNames.append('var'+str(i+1))
	
	if phjPrintResults == True:	
		# Print variable names - just for reference...
		for i in range (len(phjTempVariableNames)):
			print("var",i+1,": ",phjTempVariableNames[i])
	
	return phjTempVariableNames




####################################
# Read data from an Stata datafile #
####################################
def phjDataReadFromStata(phjStataFileName,
						 phjPrintResults = False):

# The following method to check that a package is available is not considered to be
# good practice:
#
#	try:
# 		import package_name
# 		HAVE_PACKAGE = True
# 	except ImportError:
# 		HAVE_PACKAGE = False
#
# Instead use the following method (as described at: http://developer.plone.org/reference_manuals/external/plone.api/contribute/conventions.html#about-imports)...
	
#	The Pandas I/O api is a set of top level reader functions accessed
#	like pd.read_csv() that generally return a pandas object.
#	(See: http://pandas.pydata.org/pandas-docs/stable/io.html)
#	
#	pandas.io.stata.read_stata
#	==========================
#	Taken from: http://pandas.pydata.org/pandas-docs/dev/generated/pandas.io.stata.read_stata.html
#	(Accessed 3 Feb 2014)
#	
#	pandas.io.stata.read_stata(filepath_or_buffer, convert_dates=True, convert_categoricals=True, encoding=None, index=None)
#	Read Stata file into DataFrame
#	
#	Parameters :	
#		filepath_or_buffer : string or file-like object
#			Path to .dta file or object implementing a binary read() functions
#		convert_dates : boolean, defaults to True
#			Convert date variables to DataFrame time values
#		convert_categoricals : boolean, defaults to True
#			Read value labels and convert columns to Categorical/Factor variables
#		ncoding : string, None or encoding
#			Encoding used to parse the files. Note that Stata does not support unicode. None defaults to cp1252.
#		index : identifier of index column
#			identifier of column that should be used as index of the DataFrame
	
	if phjPrintResults:
		print('\nStata filename: ',phjStataFileName)
	
	try:
		phjStataData = pd.read_stata(phjStataFileName, convert_dates=True, convert_categoricals=True, encoding=None, index=None)
	except:		# Catch all exceptions
#		e = sys.exc_info()[0]
		print("An error occurred reading data from Stata file.")
		exit(0)
	
	if phjPrintResults:
		print(phjStataData)
		
	return phjStataData



#################################
# Read data from an R workspace #
#################################
def phjReadDataFromR(phjPathToNewWorkingDirectory,
					 phjRData,
					 phjPrintResults = False):

# The following method to check that a package is available is not considered to be
# good practice:
#
#	try:
# 		import package_name
# 		HAVE_PACKAGE = True
# 	except ImportError:
# 		HAVE_PACKAGE = False
#
# Instead use the following method (as described at: http://developer.plone.org/reference_manuals/external/plone.api/contribute/conventions.html#about-imports)...
	
	import pkg_resources
			
	try:
		pkg_resources.get_distribution('pandas')
	except pkg_resources.DistributionNotFound:
		print("Pandas package not available.")
		exit(0)
	else:
		import pandas as pd
		import pandas.rpy.common as com
		
	try:
		pkg_resources.get_distribution('rpy2')
	except pkg_resources.DistributionNotFound:
		print("rpy2 package not available.")
		exit(0)
	else:
		import rpy2.robjects as robjects
		
	
	print(robjects.r("setwd('"+phjPathToNewWorkingDirectory+"')"))
	
	print(robjects.r.load(".RData"))
	
	myRData = com.load_data(phjRData)
	
	return myRData
	
	

